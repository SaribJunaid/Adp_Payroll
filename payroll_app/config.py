import streamlit as st
from openpyxl.styles import PatternFill, Border, Side

REG_RATE = 24
OT_RATE = 36

REDIRECT_URI = "https://adppayroll.streamlit.app/"
SCOPES = [
    "https://graph.microsoft.com/Files.ReadWrite.All",
    "https://graph.microsoft.com/Sites.ReadWrite.All",
    "https://graph.microsoft.com/User.Read",
]

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
RELAY_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ADP_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DRIVER_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OVERRIDE_FILL = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
ANOMALY_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def load_azure_credentials():
    try:
        azure = st.secrets["azure"]
        return {
            "client_id": azure["client_id"],
            "tenant_id": azure["tenant_id"],
            "client_secret": azure["client_secret"],
            "configured": True,
        }
    except Exception:
        return {
            "client_id": None,
            "tenant_id": None,
            "client_secret": None,
            "configured": False,
        }

