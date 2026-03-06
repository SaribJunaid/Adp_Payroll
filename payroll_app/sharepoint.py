import io
from urllib.parse import urlencode, urlparse

import msal
import requests
from openpyxl import load_workbook

from payroll_app.config import REDIRECT_URI, SCOPES


def get_auth_url(client_id, tenant_id):
    auth_endpoint = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/authorize"
    params = {
        "client_id": client_id,
        "response_type": "code",
        "redirect_uri": REDIRECT_URI,
        "scope": " ".join(SCOPES),
        "state": "payroll",
    }
    return f"{auth_endpoint}?{urlencode(params)}"


def exchange_code_for_token(auth_code, client_id, tenant_id, client_secret):
    try:
        app = msal.ConfidentialClientApplication(
            client_id,
            client_credential=client_secret,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
        )
        result = app.acquire_token_by_authorization_code(
            auth_code,
            scopes=SCOPES,
            redirect_uri=REDIRECT_URI,
        )
        return result.get("access_token"), result.get("error_description")
    except Exception as exc:
        return None, str(exc)


def get_user_info(access_token):
    resp = requests.get(
        "https://graph.microsoft.com/v1.0/me",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=30,
    )
    if resp.status_code == 200:
        return resp.json()
    return None


def get_site_from_url(access_token, sharepoint_url):
    try:
        parsed = urlparse(sharepoint_url)
        path_parts = parsed.path.strip("/").split("/")
        headers = {"Authorization": f"Bearer {access_token}"}
        if "sites" in path_parts:
            site_slug = path_parts[path_parts.index("sites") + 1]
            url = f"https://graph.microsoft.com/v1.0/sites/{parsed.hostname}:/sites/{site_slug}"
        else:
            url = f"https://graph.microsoft.com/v1.0/sites/{parsed.hostname}"
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 200:
            return resp.json(), None
        return None, f"Error: {resp.status_code}"
    except Exception as exc:
        return None, str(exc)


def list_sharepoint_files(access_token, site_id, path="root"):
    try:
        headers = {"Authorization": f"Bearer {access_token}"}
        if path == "root":
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children"
        else:
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{path}:/children"
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 200:
            return resp.json().get("value", []), None
        return None, "Folder error"
    except Exception as exc:
        return None, str(exc)


def download_sharepoint_file(access_token, site_id, file_id):
    try:
        resp = requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/content",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=90,
        )
        if resp.status_code == 200:
            return resp.content, None
        return None, "Download error"
    except Exception as exc:
        return None, str(exc)


def upload_to_sharepoint(access_token, site_id, folder_path, filename, file_content):
    try:
        if folder_path != "root":
            url = (
                f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:"
                f"/{folder_path}/{filename}:/content"
            )
        else:
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{filename}:/content"
        resp = requests.put(
            url,
            headers={"Authorization": f"Bearer {access_token}"},
            data=file_content,
            timeout=90,
        )
        if resp.status_code in [200, 201]:
            return True, None
        return False, f"Error: {resp.status_code}"
    except Exception as exc:
        return False, str(exc)


def check_workbook_exists(access_token, site_id, folder_path, workbook_name="ADP.xlsx"):
    try:
        if folder_path != "root":
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}/{workbook_name}"
        else:
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{workbook_name}"
        resp = requests.get(url, headers={"Authorization": f"Bearer {access_token}"}, timeout=30)
        if resp.status_code == 200:
            return True, resp.json().get("id")
        return False, None
    except Exception:
        return False, None


def add_sheet_to_workbook(existing_wb_bytes, new_sheet_bytes, sheet_name):
    try:
        wb = load_workbook(io.BytesIO(existing_wb_bytes))
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        new_wb = load_workbook(io.BytesIO(new_sheet_bytes))
        new_ws = new_wb.active
        for row in new_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, None
    except Exception as exc:
        return None, str(exc)

