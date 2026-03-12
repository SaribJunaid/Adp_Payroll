# import io
# from datetime import datetime
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Font, PatternFill
# from openpyxl.utils import get_column_letter
# from payroll_app.config import * # Ensure HEADER_FILL, RELAY_FILL, ADP_FILL, DRIVER_FILL, OVERRIDE_FILL, ANOMALY_FILL, THICK_BORDER, THIN_BORDER are defined

# def create_excel(final_df, relay_cols, adp_cols, override_map):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Payroll"

#     # Define the calculation columns to be added at the end
#     calc_columns = [
#         "W1_Relay_Loads", "W2_Relay_Loads", "Total_Relay_Loads",
#         "W1 Hours", "W1 Regular", "W1 OT", 
#         "W2 Hours", "W2 Regular", "W2 OT",
#         "Total ADP Hours", "Total Regular", "Total OT",
#         "Override Pay", "Final Pay", "Pay by Hours", "Hour Adjustment"
#     ]

#     # Convert date columns for display (e.g., R_2023-01-01 -> 01-Jan)
#     # We use the column names directly to ensure we stay synced with the DataFrame
#     relay_display = [datetime.strptime(c.replace("R_", ""), "%Y-%m-%d").strftime("%d-%b") for c in relay_cols]
#     adp_display = [datetime.strptime(c.replace("A_", ""), "%Y-%m-%d").strftime("%d-%b") for c in adp_cols]

#     # Combine all headers
#     base_headers = [c for c in final_df.columns if not c.startswith(("R_", "A_"))]
#     headers = base_headers + relay_display + adp_display + calc_columns
#     ws.append(headers)
    
#     # Create a mapping of Header Name -> Column Letter for formulas
#     col_map = {name: get_column_letter(idx) for idx, name in enumerate(headers, 1)}
#     def gcl(name): return col_map.get(name)

#     # Apply Header Styling
#     for cell in ws[1]:
#         cell.fill = HEADER_FILL
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = THIN_BORDER

#     # --- Helper functions for TIER_SWITCH formula building ---
#     # Defined once outside the row loop for efficiency.

#     def _tier_base(n_cell, sw_at_cell):
#         """
#         Excel nested-IF returning fixed tier pay for up to Switch_After_Load loads.
#         Tiers: 1 load=350, 2 loads=700, 3+ loads=1000.
#         n_cell is capped at sw_at_cell before lookup.
#         """
#         m = f"MIN({n_cell},{sw_at_cell})"
#         return f"IF({m}=1,350,IF({m}=2,700,IF({m}>=3,1000,0)))"

#     def _week_tier_formula(wk_lds, sw_at_cell, sw_type_cell, extra_hrs):
#         """
#         Per-week TIER_SWITCH pay:
#           loads <= switch  →  tier_base(loads)
#           loads >  switch  →  tier_base(switch) + extra
#             extra: FIXED  = (loads - switch) * 350
#                    HOURLY = pre-computed extra_hrs * 24
#         """
#         base        = _tier_base(wk_lds, sw_at_cell)
#         extra_fixed = f"({wk_lds}-{sw_at_cell})*350"
#         extra_hrly  = f"{extra_hrs}*24"
#         extra       = f"IF({sw_type_cell}=\"FIXED\",{extra_fixed},{extra_hrly})"
#         return f"IF({wk_lds}<={sw_at_cell},{base},{base}+{extra})"

#     # Iterate through Data Rows
#     for row_idx, row_values in enumerate(final_df.values, start=2):
#         row_num = str(row_idx)
#         driver_name = str(row_values[0])
        
#         # 1. Write the base data from the DataFrame
#         for c_idx, val in enumerate(row_values):
#             cell = ws.cell(row_idx, c_idx + 1, val)
            
#             # Formatting: Hide zeros for cleaner look
#             if val == 0:
#                 cell.number_format = ';;;' 
            
#             # Apply color coding based on column category
#             header = headers[c_idx]
#             if c_idx == 0: # Driver Name Column
#                 cell.fill = DRIVER_FILL
#             elif header in relay_display:
#                 cell.fill = RELAY_FILL
#             elif header in adp_display:
#                 cell.fill = ADP_FILL
            
#             cell.border = THIN_BORDER
#             cell.alignment = Alignment(horizontal="center")

#         col_ptr = len(row_values) + 1
        
#         # --- FORMULA SECTION ---
#         # 2. Weekly Relay Loads
#         # After dropping the first date, we have 14 relay columns in order.
#         # W1 = first 7 relay cols, W2 = last 7 relay cols.
#         # We use direct column index arithmetic (not name lookup) to guarantee
#         # the correct cells are always referenced regardless of date display strings.

#         # Calculate the absolute sheet column index where relay data starts
#         relay_start_col = len(base_headers) + 1  # 1-based sheet column index

#         # W1: first 7 relay columns
#         w1_start_letter = get_column_letter(relay_start_col)
#         w1_end_letter   = get_column_letter(relay_start_col + min(6, len(relay_display) - 1))
#         w1_relay_range  = f"{w1_start_letter}{row_num}:{w1_end_letter}{row_num}"
#         ws.cell(row_idx, col_ptr).value = f'=COUNTIF({w1_relay_range}, ">0")'

#         # W2: next 7 relay columns (indices 7-13)
#         w2_relay_val = 0
#         if len(relay_display) > 7:
#             w2_start_letter = get_column_letter(relay_start_col + 7)
#             w2_end_letter   = get_column_letter(relay_start_col + min(13, len(relay_display) - 1))
#             w2_relay_range  = f"{w2_start_letter}{row_num}:{w2_end_letter}{row_num}"
#             w2_relay_val    = f'=COUNTIF({w2_relay_range}, ">0")'
        
#         ws.cell(row_idx, col_ptr + 1).value = w2_relay_val
#         ws.cell(row_idx, col_ptr + 2).value = f"={gcl('W1_Relay_Loads')}{row_num}+{gcl('W2_Relay_Loads')}{row_num}"
#         col_ptr += 3

#         # 3. ADP Week 1 & 2 Calculation
#         # ADP columns start immediately after all relay columns.
#         adp_start_col = len(base_headers) + len(relay_display) + 1  # 1-based

#         # ADP Week 1: first 7 ADP columns
#         w1_adp_start_letter = get_column_letter(adp_start_col)
#         w1_adp_end_letter   = get_column_letter(adp_start_col + min(6, len(adp_display) - 1))
#         w1_adp_range = f"{w1_adp_start_letter}{row_num}:{w1_adp_end_letter}{row_num}"
#         ws.cell(row_idx, col_ptr).value = f"=SUM({w1_adp_range})"
#         ws.cell(row_idx, col_ptr + 1).value = f"=MIN({gcl('W1 Hours')}{row_num}, 40)"
#         ws.cell(row_idx, col_ptr + 2).value = f"=MAX(0, {gcl('W1 Hours')}{row_num}-40)"
#         col_ptr += 3

#         # ADP Week 2: next 7 ADP columns
#         if len(adp_display) > 7:
#             w2_adp_start_letter = get_column_letter(adp_start_col + 7)
#             w2_adp_end_letter   = get_column_letter(adp_start_col + min(13, len(adp_display) - 1))
#             w2_adp_range = f"{w2_adp_start_letter}{row_num}:{w2_adp_end_letter}{row_num}"
#             ws.cell(row_idx, col_ptr).value = f"=SUM({w2_adp_range})"
#             ws.cell(row_idx, col_ptr + 1).value = f"=MIN({gcl('W2 Hours')}{row_num}, 40)"
#             ws.cell(row_idx, col_ptr + 2).value = f"=MAX(0, {gcl('W2 Hours')}{row_num}-40)"
#         else:
#             ws.cell(row_idx, col_ptr).value = 0
#             ws.cell(row_idx, col_ptr + 1).value = 0
#             ws.cell(row_idx, col_ptr + 2).value = 0
#         col_ptr += 3

#         # 4. ADP Totals
#         ws.cell(row_idx, col_ptr).value = f"={gcl('W1 Hours')}{row_num}+{gcl('W2 Hours')}{row_num}"
#         ws.cell(row_idx, col_ptr + 1).value = f"={gcl('W1 Regular')}{row_num}+{gcl('W2 Regular')}{row_num}"
#         ws.cell(row_idx, col_ptr + 2).value = f"={gcl('W1 OT')}{row_num}+{gcl('W2 OT')}{row_num}"
#         col_ptr += 3

#         # 5. Overrides
#         # We find the override total for the specific driver
#         driver_overrides = [v for (d, dt), v in override_map.items() if d.upper() == driver_name.upper()]
#         override_total = sum(driver_overrides)
#         ov_cell = ws.cell(row_idx, col_ptr)
#         ov_cell.value = override_total
#         if override_total > 0:
#             ov_cell.fill = OVERRIDE_FILL
#         col_ptr += 1

#         # 6. Final Pay Logic
#         cat      = f"{gcl('Category')}{row_num}"
#         w1_lds   = f"{gcl('W1_Relay_Loads')}{row_num}"
#         w2_lds   = f"{gcl('W2_Relay_Loads')}{row_num}"
#         tot_lds  = f"{gcl('Total_Relay_Loads')}{row_num}"
#         pkg_amt  = f"{gcl('Package_Amount')}{row_num}"
#         hr_rate  = f"{gcl('Hourly_Rate')}{row_num}"
#         w1_adp_hr  = f"{gcl('W1 Hours')}{row_num}"
#         w2_adp_hr  = f"{gcl('W2 Hours')}{row_num}"
#         tot_adp_hr = f"{gcl('Total ADP Hours')}{row_num}"
#         sw_at    = f"{gcl('Switch_After_Load')}{row_num}"
#         sw_type  = f"{gcl('Switch_Type')}{row_num}"   # "FIXED" or "HOURLY"
#         target   = f"{gcl('Target_Load')}{row_num}"
#         ov_pay   = f"{get_column_letter(col_ptr-1)}{row_num}"

#         # --- TIER_SWITCH base pay helper (per week) ---
#         # Fixed tier values: 1 load=350, 2 loads=700, 3 loads=1000
#         # For loads > Switch_After_Load the base is capped at the Switch tier value.
#         # Extra loads beyond switch: FIXED=+350/load, HOURLY=extra_load_ADP_hrs×24

#         # Python-side: compute extra-load ADP hours for W1 and W2.
#         row_series = dict(zip(final_df.columns, row_values))

#         # Compute extra-load ADP hours in Python per week.
#         # Relay days with hours > 0 are real loads, sorted chronologically.
#         # Extra loads = those beyond Switch_After_Load threshold.
#         # We look up the ADP hours for the same calendar date as each extra relay day.
#         def _extra_adp_hrs(relay_week_cols, adp_week_cols):
#             try:
#                 sw_val = int(row_series.get("Switch_After_Load", 0) or 0)
#             except (ValueError, TypeError):
#                 sw_val = 0
#             relay_vals = [(c, float(row_series.get(c, 0) or 0)) for c in relay_week_cols]
#             load_days  = [c for c, v in relay_vals if v > 0]
#             if len(load_days) <= sw_val:
#                 return 0.0
#             extra_days   = load_days[sw_val:]
#             adp_date_map = {c.replace("A_", ""): c for c in adp_week_cols}
#             total = 0.0
#             for r_col in extra_days:
#                 a_col = adp_date_map.get(r_col.replace("R_", ""))
#                 if a_col:
#                     total += float(row_series.get(a_col, 0) or 0)
#             return round(total, 4)

#         w1_relay_cols = relay_cols[:7]
#         w2_relay_cols = relay_cols[7:14]
#         w1_adp_cols   = adp_cols[:7]
#         w2_adp_cols   = adp_cols[7:14]

#         w1_extra_hrs = _extra_adp_hrs(w1_relay_cols, w1_adp_cols)
#         w2_extra_hrs = _extra_adp_hrs(w2_relay_cols, w2_adp_cols)

#         w1_tier = _week_tier_formula(w1_lds, sw_at, sw_type, w1_extra_hrs)
#         w2_tier = _week_tier_formula(w2_lds, sw_at, sw_type, w2_extra_hrs)

#         # Full final pay formula.
#         # Drivers NOT in DriverPay will have no Category (blank/NaN) —
#         # for them Final Pay = Total Regular×24 + Total OT×36 (pure hours-based).
#         tot_reg = f"{gcl('Total Regular')}{row_num}"
#         tot_ot  = f"{gcl('Total OT')}{row_num}"

#         # Detect whether this driver has a Category (i.e. exists in DriverPay)
#         driver_category = str(row_series.get("Category", "")).strip()
#         driver_in_driverpay = driver_category != "" and driver_category.lower() != "nan"

#         if driver_in_driverpay:
#             formula = (
#                 f"=IF({cat}=\"PER_LOAD\",{tot_lds}*350,"
#                 f"IF({cat}=\"TARGET\",IF({tot_lds}>={target},{pkg_amt},{tot_lds}*350),"
#                 f"IF({cat}=\"TIER_SWITCH\","
#                 f"({w1_tier})+({w2_tier}),"
#                 f"{tot_adp_hr}*24)))"
#                 f"+{ov_pay}"
#             )
#         else:
#             # Not in DriverPay — pay purely by ADP hours
#             formula = f"={tot_reg}*24+{tot_ot}*36"

#         ws.cell(row_idx, col_ptr).value = formula
#         final_pay_col = get_column_letter(col_ptr)
#         col_ptr += 1

#         # 7. Pay by Hours: Total Regular × 24 + Total OT × 36
#         pay_by_hours_col = get_column_letter(col_ptr)
#         ws.cell(row_idx, col_ptr).value = f"={tot_reg}*24+{tot_ot}*36"
#         col_ptr += 1

#         # 8. Hour Adjustment: (Final Pay - Pay by Hours) / 36
#         # = how many extra hours (each worth $36) close the gap.
#         # Hidden (blank) for drivers not in DriverPay since their gap is always 0.
#         if driver_in_driverpay:
#             ws.cell(row_idx, col_ptr).value = (
#                 f"=({final_pay_col}{row_num}-{pay_by_hours_col}{row_num})/36"
#             )
#         else:
#             ws.cell(row_idx, col_ptr).value = None  # blank — no adjustment needed

#         # --- ANOMALY DETECTION ---
#         # If Relay hours exist but ADP hours are 0, highlight driver name
#         relay_start_idx = len(base_headers)
#         relay_end_idx = relay_start_idx + len(relay_cols)
#         adp_start_idx = relay_end_idx
#         adp_end_idx = adp_start_idx + len(adp_cols)
        
#         relay_sum = sum([v for v in row_values[relay_start_idx:relay_end_idx] if isinstance(v, (int, float))])
#         adp_sum = sum([v for v in row_values[adp_start_idx:adp_end_idx] if isinstance(v, (int, float))])
        
#         if relay_sum > 0 and adp_sum == 0:
#             ws.cell(row_idx, 1).fill = ANOMALY_FILL

#         # Formatting for the added calculation columns
#         for c in range(len(row_values) + 1, col_ptr + 1):
#             cell = ws.cell(row_idx, c)
#             cell.border = THIN_BORDER
#             cell.alignment = Alignment(horizontal="center")
#             if cell.value == 0:
#                 cell.number_format = ';;;'

#     # Auto-adjust column widths for readability
#     for col in ws.columns:
#         ws.column_dimensions[col[0].column_letter].width = 16

#     out = io.BytesIO()
#     wb.save(out)
#     out.seek(0)
#     return out, []

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from payroll_app.config import * # Ensure HEADER_FILL, RELAY_FILL, ADP_FILL, DRIVER_FILL, OVERRIDE_FILL, ANOMALY_FILL, THICK_BORDER, THIN_BORDER are defined

def create_excel(final_df, relay_cols, adp_cols, override_map):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"

    # Define the calculation columns to be added at the end
    calc_columns = [
        "W1_Relay_Loads", "W2_Relay_Loads", "Total_Relay_Loads",
        "Solo5_Count", "Solo5_Pay",
        "W1 Hours", "W1 Regular", "W1 OT", 
        "W2 Hours", "W2 Regular", "W2 OT",
        "Total ADP Hours", "Total Regular", "Total OT",
        "Override Pay", "Final Pay", "Pay by Hours", "Hour Adjustment"
    ]

    # Convert date columns for display (e.g., R_2023-01-01 -> 01-Jan)
    # We use the column names directly to ensure we stay synced with the DataFrame
    relay_display = [datetime.strptime(c.replace("R_", ""), "%Y-%m-%d").strftime("%d-%b") for c in relay_cols]
    adp_display = [datetime.strptime(c.replace("A_", ""), "%Y-%m-%d").strftime("%d-%b") for c in adp_cols]

    # Combine all headers
    base_headers = [c for c in final_df.columns if not c.startswith(("R_", "A_"))]
    headers = base_headers + relay_display + adp_display + calc_columns
    ws.append(headers)
    
    # Create a mapping of Header Name -> Column Letter for formulas
    col_map = {name: get_column_letter(idx) for idx, name in enumerate(headers, 1)}
    def gcl(name): return col_map.get(name)

    # Apply Header Styling
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # --- Helper functions for TIER_SWITCH formula building ---
    # Defined once outside the row loop for efficiency.

    def _tier_base(n_cell, sw_at_cell):
        """
        Excel nested-IF returning fixed tier pay for up to Switch_After_Load loads.
        Tiers: 1 load=350, 2 loads=700, 3+ loads=1000.
        n_cell is capped at sw_at_cell before lookup.
        """
        m = f"MIN({n_cell},{sw_at_cell})"
        return f"IF({m}=1,350,IF({m}=2,700,IF({m}>=3,1000,0)))"

    def _week_tier_formula(wk_lds, sw_at_cell, sw_type_cell, extra_hrs):
        """
        Per-week TIER_SWITCH pay:
          loads <= switch  →  tier_base(loads)
          loads >  switch  →  tier_base(switch) + extra
            extra: FIXED  = (loads - switch) * 350
                   HOURLY = pre-computed extra_hrs * 24
        """
        base        = _tier_base(wk_lds, sw_at_cell)
        extra_fixed = f"({wk_lds}-{sw_at_cell})*350"
        extra_hrly  = f"{extra_hrs}*24"
        extra       = f"IF({sw_type_cell}=\"FIXED\",{extra_fixed},{extra_hrly})"
        return f"IF({wk_lds}<={sw_at_cell},{base},{base}+{extra})"

    # Iterate through Data Rows
    for row_idx, row_values in enumerate(final_df.values, start=2):
        row_num = str(row_idx)
        driver_name = str(row_values[0])
        
        # 1. Write the base data from the DataFrame
        for c_idx, val in enumerate(row_values):
            cell = ws.cell(row_idx, c_idx + 1, val)
            
            # Formatting: Hide zeros for cleaner look
            if val == 0:
                cell.number_format = ';;;' 
            
            # Apply color coding based on column category
            header = headers[c_idx]
            if c_idx == 0: # Driver Name Column
                cell.fill = DRIVER_FILL
            elif header in relay_display:
                cell.fill = RELAY_FILL
            elif header in adp_display:
                cell.fill = ADP_FILL
            
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        col_ptr = len(row_values) + 1
        
        # --- FORMULA SECTION ---
        # 2. Weekly Relay Loads
        # After dropping the first date, we have 14 relay columns in order.
        # W1 = first 7 relay cols, W2 = last 7 relay cols.
        # We use direct column index arithmetic (not name lookup) to guarantee
        # the correct cells are always referenced regardless of date display strings.

        # Calculate the absolute sheet column index where relay data starts
        relay_start_col = len(base_headers) + 1  # 1-based sheet column index

        # W1: first 7 relay columns
        w1_start_letter = get_column_letter(relay_start_col)
        w1_end_letter   = get_column_letter(relay_start_col + min(6, len(relay_display) - 1))
        w1_relay_range  = f"{w1_start_letter}{row_num}:{w1_end_letter}{row_num}"
        ws.cell(row_idx, col_ptr).value = f'=COUNTIF({w1_relay_range}, ">0")'

        # W2: next 7 relay columns (indices 7-13)
        w2_relay_val = 0
        if len(relay_display) > 7:
            w2_start_letter = get_column_letter(relay_start_col + 7)
            w2_end_letter   = get_column_letter(relay_start_col + min(13, len(relay_display) - 1))
            w2_relay_range  = f"{w2_start_letter}{row_num}:{w2_end_letter}{row_num}"
            w2_relay_val    = f'=COUNTIF({w2_relay_range}, ">0")'
        
        ws.cell(row_idx, col_ptr + 1).value = w2_relay_val
        ws.cell(row_idx, col_ptr + 2).value = f"={gcl('W1_Relay_Loads')}{row_num}+{gcl('W2_Relay_Loads')}{row_num}"
        col_ptr += 3

        # Build row_series once here — used by Solo5, extra ADP hours, and category detection
        row_series = dict(zip(final_df.columns, row_values))

        # Solo5 Detection (Python-side):
        # Any relay date column where hours > 30 = one Solo5 block = $1800 flat.
        # Multiple Solo5 blocks in the same period each add $1800.
        solo5_count = sum(
            1 for c in relay_cols
            if float(row_series.get(c, 0) or 0) > 30
        )
        solo5_pay = solo5_count * 1800

        # Write Solo5_Count and Solo5_Pay cells — colored purple when present
        SOLO5_FILL = PatternFill("solid", fgColor="9B59B6")  # purple

        solo5_count_cell = ws.cell(row_idx, col_ptr)
        solo5_pay_cell   = ws.cell(row_idx, col_ptr + 1)

        if solo5_count > 0:
            solo5_count_cell.value = solo5_count
            solo5_count_cell.fill  = SOLO5_FILL
            solo5_count_cell.font  = Font(bold=True, color="FFFFFF")
            solo5_pay_cell.value   = solo5_pay
            solo5_pay_cell.fill    = SOLO5_FILL
            solo5_pay_cell.font    = Font(bold=True, color="FFFFFF")

        # Also highlight the relay date cells that triggered Solo5 (hours > 30)
        relay_start_idx = len(base_headers)
        for i, rc in enumerate(relay_cols):
            if float(row_series.get(rc, 0) or 0) > 30:
                relay_cell = ws.cell(row_idx, relay_start_idx + i + 1)
                relay_cell.fill = SOLO5_FILL
                relay_cell.font = Font(bold=True, color="FFFFFF")

        col_ptr += 2
        # ADP columns start immediately after all relay columns.
        adp_start_col = len(base_headers) + len(relay_display) + 1  # 1-based

        # ADP Week 1: first 7 ADP columns
        w1_adp_start_letter = get_column_letter(adp_start_col)
        w1_adp_end_letter   = get_column_letter(adp_start_col + min(6, len(adp_display) - 1))
        w1_adp_range = f"{w1_adp_start_letter}{row_num}:{w1_adp_end_letter}{row_num}"
        ws.cell(row_idx, col_ptr).value = f"=SUM({w1_adp_range})"
        ws.cell(row_idx, col_ptr + 1).value = f"=MIN({gcl('W1 Hours')}{row_num}, 40)"
        ws.cell(row_idx, col_ptr + 2).value = f"=MAX(0, {gcl('W1 Hours')}{row_num}-40)"
        col_ptr += 3

        # ADP Week 2: next 7 ADP columns
        if len(adp_display) > 7:
            w2_adp_start_letter = get_column_letter(adp_start_col + 7)
            w2_adp_end_letter   = get_column_letter(adp_start_col + min(13, len(adp_display) - 1))
            w2_adp_range = f"{w2_adp_start_letter}{row_num}:{w2_adp_end_letter}{row_num}"
            ws.cell(row_idx, col_ptr).value = f"=SUM({w2_adp_range})"
            ws.cell(row_idx, col_ptr + 1).value = f"=MIN({gcl('W2 Hours')}{row_num}, 40)"
            ws.cell(row_idx, col_ptr + 2).value = f"=MAX(0, {gcl('W2 Hours')}{row_num}-40)"
        else:
            ws.cell(row_idx, col_ptr).value = 0
            ws.cell(row_idx, col_ptr + 1).value = 0
            ws.cell(row_idx, col_ptr + 2).value = 0
        col_ptr += 3

        # 4. ADP Totals
        ws.cell(row_idx, col_ptr).value = f"={gcl('W1 Hours')}{row_num}+{gcl('W2 Hours')}{row_num}"
        ws.cell(row_idx, col_ptr + 1).value = f"={gcl('W1 Regular')}{row_num}+{gcl('W2 Regular')}{row_num}"
        ws.cell(row_idx, col_ptr + 2).value = f"={gcl('W1 OT')}{row_num}+{gcl('W2 OT')}{row_num}"
        col_ptr += 3

        # 5. Overrides
        # We find the override total for the specific driver
        driver_overrides = [v for (d, dt), v in override_map.items() if d.upper() == driver_name.upper()]
        override_total = sum(driver_overrides)
        ov_cell = ws.cell(row_idx, col_ptr)
        ov_cell.value = override_total
        if override_total > 0:
            ov_cell.fill = OVERRIDE_FILL
        col_ptr += 1

        # 6. Final Pay Logic
        cat      = f"{gcl('Category')}{row_num}"
        w1_lds   = f"{gcl('W1_Relay_Loads')}{row_num}"
        w2_lds   = f"{gcl('W2_Relay_Loads')}{row_num}"
        tot_lds  = f"{gcl('Total_Relay_Loads')}{row_num}"
        pkg_amt  = f"{gcl('Package_Amount')}{row_num}"
        hr_rate  = f"{gcl('Hourly_Rate')}{row_num}"
        w1_adp_hr  = f"{gcl('W1 Hours')}{row_num}"
        w2_adp_hr  = f"{gcl('W2 Hours')}{row_num}"
        tot_adp_hr = f"{gcl('Total ADP Hours')}{row_num}"
        sw_at    = f"{gcl('Switch_After_Load')}{row_num}"
        sw_type  = f"{gcl('Switch_Type')}{row_num}"   # "FIXED" or "HOURLY"
        target   = f"{gcl('Target_Load')}{row_num}"
        ov_pay   = f"{get_column_letter(col_ptr-1)}{row_num}"

        # --- TIER_SWITCH base pay helper (per week) ---
        # Fixed tier values: 1 load=350, 2 loads=700, 3 loads=1000
        # For loads > Switch_After_Load the base is capped at the Switch tier value.
        # Extra loads beyond switch: FIXED=+350/load, HOURLY=extra_load_ADP_hrs×24

        # Compute extra-load ADP hours in Python per week.
        # Relay days with hours > 0 are real loads, sorted chronologically.
        # Extra loads = those beyond Switch_After_Load threshold.
        # We look up the ADP hours for the same calendar date as each extra relay day.
        def _extra_adp_hrs(relay_week_cols, adp_week_cols):
            try:
                sw_val = int(row_series.get("Switch_After_Load", 0) or 0)
            except (ValueError, TypeError):
                sw_val = 0
            relay_vals = [(c, float(row_series.get(c, 0) or 0)) for c in relay_week_cols]
            load_days  = [c for c, v in relay_vals if v > 0]
            if len(load_days) <= sw_val:
                return 0.0
            extra_days   = load_days[sw_val:]
            adp_date_map = {c.replace("A_", ""): c for c in adp_week_cols}
            total = 0.0
            for r_col in extra_days:
                a_col = adp_date_map.get(r_col.replace("R_", ""))
                if a_col:
                    total += float(row_series.get(a_col, 0) or 0)
            return round(total, 4)

        w1_relay_cols = relay_cols[:7]
        w2_relay_cols = relay_cols[7:14]
        w1_adp_cols   = adp_cols[:7]
        w2_adp_cols   = adp_cols[7:14]

        w1_extra_hrs = _extra_adp_hrs(w1_relay_cols, w1_adp_cols)
        w2_extra_hrs = _extra_adp_hrs(w2_relay_cols, w2_adp_cols)

        w1_tier = _week_tier_formula(w1_lds, sw_at, sw_type, w1_extra_hrs)
        w2_tier = _week_tier_formula(w2_lds, sw_at, sw_type, w2_extra_hrs)

        # Full final pay formula.
        # Drivers NOT in DriverPay will have no Category (blank/NaN) —
        # for them Final Pay = Total Regular×24 + Total OT×36 (pure hours-based).
        tot_reg = f"{gcl('Total Regular')}{row_num}"
        tot_ot  = f"{gcl('Total OT')}{row_num}"

        # Detect whether this driver has a Category (i.e. exists in DriverPay)
        driver_category = str(row_series.get("Category", "")).strip()
        driver_in_driverpay = driver_category != "" and driver_category.lower() != "nan"

        if driver_in_driverpay:
            formula = (
                f"=IF({cat}=\"PER_LOAD\",{tot_lds}*350,"
                f"IF({cat}=\"TARGET\",IF({tot_lds}>={target},{pkg_amt},{tot_lds}*350),"
                f"IF({cat}=\"TIER_SWITCH\","
                f"({w1_tier})+({w2_tier}),"
                f"{tot_adp_hr}*24)))"
                f"+{ov_pay}"
                f"+{solo5_pay}"  # Solo5 flat pay added on top regardless of category
            )
        else:
            # Not in DriverPay — pay purely by ADP hours + any Solo5
            formula = f"={tot_reg}*24+{tot_ot}*36+{solo5_pay}"

        ws.cell(row_idx, col_ptr).value = formula
        final_pay_col = get_column_letter(col_ptr)
        col_ptr += 1

        # 7. Pay by Hours: Total Regular × 24 + Total OT × 36
        pay_by_hours_col = get_column_letter(col_ptr)
        ws.cell(row_idx, col_ptr).value = f"={tot_reg}*24+{tot_ot}*36"
        col_ptr += 1

        # 8. Hour Adjustment: (Final Pay - Pay by Hours) / 36
        # = how many extra hours (each worth $36) close the gap.
        # Hidden (blank) for drivers not in DriverPay since their gap is always 0.
        if driver_in_driverpay:
            ws.cell(row_idx, col_ptr).value = (
                f"=({final_pay_col}{row_num}-{pay_by_hours_col}{row_num})/36"
            )
        else:
            ws.cell(row_idx, col_ptr).value = None  # blank — no adjustment needed

        # --- ANOMALY DETECTION ---
        # If Relay hours exist but ADP hours are 0, highlight driver name
        relay_start_idx = len(base_headers)
        relay_end_idx = relay_start_idx + len(relay_cols)
        adp_start_idx = relay_end_idx
        adp_end_idx = adp_start_idx + len(adp_cols)
        
        relay_sum = sum([v for v in row_values[relay_start_idx:relay_end_idx] if isinstance(v, (int, float))])
        adp_sum = sum([v for v in row_values[adp_start_idx:adp_end_idx] if isinstance(v, (int, float))])
        
        if relay_sum > 0 and adp_sum == 0:
            ws.cell(row_idx, 1).fill = ANOMALY_FILL

        # Formatting for the added calculation columns
        for c in range(len(row_values) + 1, col_ptr + 1):
            cell = ws.cell(row_idx, c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if cell.value == 0:
                cell.number_format = ';;;'

    # Auto-adjust column widths for readability
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, []