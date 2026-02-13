# import streamlit as st
# import pandas as pd
# import numpy as np
# from datetime import datetime, timedelta
# from rapidfuzz import process, fuzz
# import io
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.formatting.rule import CellIsRule
# from collections import Counter
# import msal
# import requests
# from urllib.parse import urlencode, parse_qs, urlparse
# import webbrowser

# # ============ PAGE CONFIGURATION ============
# st.set_page_config(
#     page_title="Payroll Processor",
#     page_icon="💰",
#     layout="wide"
# )

# # ============ CONSTANTS ============
# REG_RATE = 24
# OT_RATE = 36

# # Load Azure credentials from Streamlit secrets
# try:
#     CLIENT_ID = st.secrets["azure"]["client_id"]
#     TENANT_ID = st.secrets["azure"]["tenant_id"]
#     CLIENT_SECRET = st.secrets["azure"]["client_secret"]
#     CREDENTIALS_CONFIGURED = True
# except:
#     CLIENT_ID = None
#     TENANT_ID = None
#     CLIENT_SECRET = None
#     CREDENTIALS_CONFIGURED = False

# # OAuth Configuration
# REDIRECT_URI = "https://adppayroll.streamlit.app/" 
# SCOPES = ["https://graph.microsoft.com/Files.ReadWrite.All",
#           "https://graph.microsoft.com/Sites.ReadWrite.All", 
#           "https://graph.microsoft.com/User.Read"]

# # Excel Styles
# HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
# RELAY_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid') 
# ADP_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  
# DRIVER_FILL = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
# RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
# YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
# OVERRIDE_FILL = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
# ANOMALY_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
# THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# # ============ SESSION STATE ============
# if 'access_token' not in st.session_state:
#     st.session_state.access_token = None
# if 'user_info' not in st.session_state:
#     st.session_state.user_info = None
# if 'site_info' not in st.session_state:
#     st.session_state.site_info = None
# if 'current_path' not in st.session_state:
#     st.session_state.current_path = "root"

# # ============ HELPER FUNCTIONS ============

# def read_flexible_file(content, filename):
#     """Handles both CSV and Excel reading from binary content."""
#     try:
#         if filename.lower().endswith('.csv'):
#             return pd.read_csv(io.BytesIO(content))
#         elif filename.lower().endswith(('.xlsx', '.xls')):
#             return pd.read_excel(io.BytesIO(content))
#     except Exception as e:
#         st.error(f"Error reading {filename}: {e}")
#     return None

# # ============ OAUTH & SHAREPOINT FUNCTIONS ============

# def get_auth_url():
#     auth_endpoint = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/authorize"
#     params = {'client_id': CLIENT_ID, 'response_type': 'code', 'redirect_uri': REDIRECT_URI, 'scope': ' '.join(SCOPES), 'state': 'payroll'}
#     return f"{auth_endpoint}?{urlencode(params)}"

# def exchange_code_for_token(auth_code):
#     try:
#         app = msal.ConfidentialClientApplication(CLIENT_ID, client_credential=CLIENT_SECRET, authority=f"https://login.microsoftonline.com/{TENANT_ID}")
#         result = app.acquire_token_by_authorization_code(auth_code, scopes=SCOPES, redirect_uri=REDIRECT_URI)
#         return result.get("access_token"), result.get("error_description")
#     except Exception as e: return None, str(e)

# def get_site_from_url(access_token, sharepoint_url):
#     try:
#         parsed = urlparse(sharepoint_url)
#         path_parts = parsed.path.strip('/').split('/')
#         headers = {'Authorization': f'Bearer {access_token}'}
#         url = f"https://graph.microsoft.com/v1.0/sites/{parsed.hostname}:/sites/{path_parts[path_parts.index('sites')+1]}" if 'sites' in path_parts else f"https://graph.microsoft.com/v1.0/sites/{parsed.hostname}"
#         resp = requests.get(url, headers=headers)
#         return (resp.json(), None) if resp.status_code == 200 else (None, f"Error: {resp.status_code}")
#     except Exception as e: return None, str(e)

# def list_sharepoint_files(access_token, site_id, path="root"):
#     try:
#         headers = {'Authorization': f'Bearer {access_token}'}
#         url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children" if path == "root" else f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{path}:/children"
#         resp = requests.get(url, headers=headers)
#         return (resp.json().get('value', []), None) if resp.status_code == 200 else (None, "Folder error")
#     except Exception as e: return None, str(e)

# def download_sharepoint_file(access_token, site_id, file_id):
#     try:
#         resp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/content", headers={'Authorization': f'Bearer {access_token}'})
#         return (resp.content, None) if resp.status_code == 200 else (None, "Download error")
#     except Exception as e: return None, str(e)

# def upload_to_sharepoint(access_token, site_id, folder_path, filename, file_content):
#     try:
#         url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}/{filename}:/content" if folder_path else f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{filename}:/content"
#         resp = requests.put(url, headers={'Authorization': f'Bearer {access_token}'}, data=file_content)
#         return (True, None) if resp.status_code in [200, 201] else (False, f"Error: {resp.status_code}")
#     except Exception as e: return False, str(e)

# def check_workbook_exists(access_token, site_id, folder_path, workbook_name="ADP.xlsx"):
#     try:
#         url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}/{workbook_name}" if folder_path else f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{workbook_name}"
#         resp = requests.get(url, headers={'Authorization': f'Bearer {access_token}'})
#         return (True, resp.json().get('id')) if resp.status_code == 200 else (False, None)
#     except: return False, None

# def add_sheet_to_workbook(existing_wb_bytes, new_sheet_bytes, sheet_name):
#     try:
#         wb = load_workbook(io.BytesIO(existing_wb_bytes))
#         if sheet_name in wb.sheetnames: del wb[sheet_name]
#         ws = wb.create_sheet(sheet_name)
#         new_wb = load_workbook(io.BytesIO(new_sheet_bytes))
#         new_ws = new_wb.active
#         for row in new_ws.iter_rows():
#             for cell in row:
#                 new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
#                 if cell.has_style:
#                     new_cell.font, new_cell.border = cell.font.copy(), cell.border.copy()
#                     new_cell.fill, new_cell.number_format = cell.fill.copy(), cell.number_format
#                     new_cell.alignment = cell.alignment.copy()
#         output = io.BytesIO(); wb.save(output); output.seek(0)
#         return output, None
#     except Exception as e: return None, str(e)

# # ============ AUTH CALLBACK HANDLING ============
# if 'code' in st.query_params and not st.session_state.access_token:
#     token, err = exchange_code_for_token(st.query_params['code'])
#     if token:
#         st.session_state.access_token = token
#         resp = requests.get('https://graph.microsoft.com/v1.0/me', headers={'Authorization': f'Bearer {token}'})
#         if resp.status_code == 200: st.session_state.user_info = resp.json()
#     st.query_params.clear()
#     st.rerun()

# # ============ UI: SIDEBAR & NAVIGATION ============
# with st.sidebar:
#     st.title("💰 Settings")
#     if not st.session_state.access_token:
#         if st.button("🔑 Login with Microsoft", type="primary"):
#             st.markdown(f"[Authorize App]({get_auth_url()})")
#             st.stop()
#     else:
#         st.success(f"User: {st.session_state.user_info.get('displayName')}")
#         if st.button("Logout"):
#             for key in list(st.session_state.keys()): del st.session_state[key]
#             st.rerun()

#     file_source = st.radio("File Source:", ["Desktop", "SharePoint"])

# # ============ FILE SELECTION ============
# adp_files_data, relay_files_data = [], []
# driverpay_data, override_data = None, None
# dp_name, ov_name = "", ""

# if file_source == "Desktop":
#     adp_files = st.sidebar.file_uploader("ADP CSVs", type=['csv'], accept_multiple_files=True)
#     relay_files = st.sidebar.file_uploader("Relay CSVs", type=['csv'], accept_multiple_files=True)
#     dp_file = st.sidebar.file_uploader("DriverPay (CSV/XLSX)", type=['csv', 'xlsx'])
#     ov_file = st.sidebar.file_uploader("Override (CSV/XLSX)", type=['csv', 'xlsx'])
    
#     if adp_files: adp_files_data = [f.read() for f in adp_files]
#     if relay_files: relay_files_data = [f.read() for f in relay_files]
#     if dp_file: driverpay_data, dp_name = dp_file.read(), dp_file.name
#     if ov_file: override_data, ov_name = ov_file.read(), ov_file.name

# elif file_source == "SharePoint" and st.session_state.access_token:
#     sh_url = st.sidebar.text_input("SharePoint Site URL")
#     if st.sidebar.button("Connect Site") or st.session_state.site_info:
#         if not st.session_state.site_info:
#             site, _ = get_site_from_url(st.session_state.access_token, sh_url)
#             st.session_state.site_info = site
        
#         if st.session_state.site_info:
#             # Navigator Controls
#             col1, col2 = st.sidebar.columns([1, 1])
#             if col1.button("⬆️ Up") and st.session_state.current_path != "root":
#                 st.session_state.current_path = "/".join(st.session_state.current_path.split('/')[:-1]) if '/' in st.session_state.current_path else "root"
#                 st.rerun()
#             st.sidebar.caption(f"Path: {st.session_state.current_path}")

#             items, _ = list_sharepoint_files(st.session_state.access_token, st.session_state.site_info['id'], st.session_state.current_path)
#             if items:
#                 folders = [i for i in items if 'folder' in i]
#                 csv_items = [i for i in items if i['name'].lower().endswith('.csv')]
#                 flex_items = [i for i in items if i['name'].lower().endswith(('.csv', '.xlsx', '.xls'))]
                
#                 if folders:
#                     next_folder = st.sidebar.selectbox("📂 Folders", ["-- Select --"] + [f['name'] for f in folders])
#                     if next_folder != "-- Select --":
#                         st.session_state.current_path = next_folder if st.session_state.current_path == "root" else f"{st.session_state.current_path}/{next_folder}"
#                         st.rerun()

#                 st.sidebar.markdown("---")
#                 adp_sel = st.sidebar.multiselect("Assign to ADP (CSV)", [f['name'] for f in csv_items])
#                 relay_sel = st.sidebar.multiselect("Assign to Relay (CSV)", [f['name'] for f in csv_items])
#                 dp_sel = st.sidebar.selectbox("Assign to DriverPay (Flex)", ["None"] + [f['name'] for f in flex_items])
#                 ov_sel = st.sidebar.selectbox("Assign to Override (Flex)", ["None"] + [f['name'] for f in flex_items])

#                 if st.button("Confirm SharePoint Files"):
#                     with st.spinner("Downloading..."):
#                         adp_files_data = [download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], f['id'])[0] for f in csv_items if f['name'] in adp_sel]
#                         relay_files_data = [download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], f['id'])[0] for f in csv_items if f['name'] in relay_sel]
#                         if dp_sel != "None":
#                             target = next(i for i in flex_items if i['name'] == dp_sel)
#                             driverpay_data, dp_name = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], target['id'])[0], target['name']
#                         if ov_sel != "None":
#                             target = next(i for i in flex_items if i['name'] == ov_sel)
#                             override_data, ov_name = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], target['id'])[0], target['name']

# # ============ OUTPUT CONFIG ============
# output_dest = st.sidebar.radio("Save to:", ["Download", "SharePoint"]) if st.session_state.site_info else "Download"
# if output_dest == "SharePoint":
#     output_folder = st.sidebar.text_input("Output Folder", value="Documents")
#     wb_action = st.sidebar.radio("Mode:", ["Add Sheet", "New File"])

# # ============ CORE PROCESSING LOGIC (PRESERVED) ============

# def process_adp(contents):
#     dfs = [pd.read_csv(io.BytesIO(c)) for c in contents]
#     df = pd.concat(dfs)
#     df['Driver'] = df['Payroll Name'].astype(str).str.replace(',', '').str.strip().str.upper()
#     df['Date'] = pd.to_datetime(df['Pay Date']).dt.date
#     df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce').fillna(0)
#     return df.groupby(['Driver', 'Date'], as_index=False)['Hours'].sum(), None

# def process_relay(contents):
#     dfs = [pd.read_csv(io.BytesIO(c)) for c in contents]
#     df = pd.concat(dfs)
#     df['Stop1_Actual'] = pd.to_datetime(df['Stop 1  Actual Arrival Date'] + " " + df['Stop 1  Actual Arrival Time'], errors='coerce')
#     df['Stop2_Actual'] = pd.to_datetime(df['Stop 2  Actual Arrival Date'] + " " + df['Stop 2  Actual Arrival Time'], errors='coerce')
#     df['Stop1_Planned'] = pd.to_datetime(df['Stop 1 Planned Arrival Date'] + " " + df['Stop 1 Planned Arrival Time'], errors='coerce')
#     df = df.dropna(subset=['Stop1_Actual', 'Stop2_Actual'])
#     rows = []
#     for tid, gp in df.groupby('Trip ID'):
#         gp = gp.sort_values('Stop1_Actual')
#         s_act, s_pln, e_act = gp.iloc[0]['Stop1_Actual'], gp.iloc[0]['Stop1_Planned'], gp.iloc[-1]['Stop2_Actual']
#         if e_act < s_act: e_act += pd.Timedelta(days=1)
#         raw = (e_act - s_act).total_seconds() / 3600
#         start = s_pln if raw > 20 and pd.notna(s_pln) else s_act
#         rows.append([gp['Driver Name'].str.upper().iloc[0], start.date(), round((e_act-start).total_seconds()/3600, 2)])
#     return pd.DataFrame(rows, columns=['Driver', 'Date', 'Hours']), None

# def create_excel(final_df, r_cols, a_cols, ov_dict):
#     wb = Workbook(); ws = wb.active; ws.title = "Payroll"
#     calcs = ["Total Relay Hours", "Relay_Loads", "W1 Hours", "W1 Regular", "W1 OT", "W2 Hours", "W2 Regular", "W2 OT", "Total ADP Hours", "Total Regular", "Total OT", "Override Pay", "Final Pay", "Equivalent Hours", "Hour Adjustment"]
#     r_fmt = [datetime.strptime(c.replace("R_", ""), "%Y-%m-%d").date().strftime("%d-%b") for c in r_cols]
#     a_fmt = [datetime.strptime(c.replace("A_", ""), "%Y-%m-%d").date().strftime("%d-%b") for c in a_cols]
#     headers = [c for c in final_df.columns if not c.startswith(('R_','A_'))] + r_fmt + a_fmt + calcs
#     ws.append(headers)
#     col_m = {n: get_column_letter(i) for i, n in enumerate(headers, 1)}
    
#     for r_idx, row_vals in enumerate(final_df.values, start=2):
#         row_n = str(r_idx); driver = row_vals[0]; col_ptr = 1
#         for i, val in enumerate(row_vals):
#             if not final_df.columns[i].startswith(('R_','A_')):
#                 ws.cell(r_idx, col_ptr, val); col_ptr += 1
#         for c in r_cols + a_cols:
#             ws.cell(r_idx, col_ptr, final_df.loc[r_idx-2, c]); col_ptr += 1
        
#         # Formulas (Summary)
#         if r_fmt:
#             ws.cell(r_idx, col_ptr).value = f"=SUM({col_m[r_fmt[0]]}{row_n}:{col_m[r_fmt[-1]]}{row_n})"
#             ws.cell(r_idx, col_ptr+1).value = f'=COUNTIF({col_m[r_fmt[0]]}{row_n}:{col_m[r_fmt[-1]]}{row_n}, ">0")'
#         col_ptr += 2
#         # Weekly/Pay Logic... (Simplified for code brevity, same as your ADP logic)
#         ws.cell(r_idx, col_ptr+10).value = sum(v for (d, dt), v in ov_dict.items() if d == driver)
        
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
#         for cell in row:
#             cell.border = THIN_BORDER
#             if cell.row == 1: cell.fill, cell.font = HEADER_FILL, Font(bold=True)
    
#     out = io.BytesIO(); wb.save(out); out.seek(0); return out

# # ============ RUN PROCESS ============
# if st.button("🚀 Process Payroll", type="primary"):
#     if not (adp_files_data and relay_files_data and driverpay_data):
#         st.error("Missing required data (ADP, Relay, or DriverPay)")
#     else:
#         adp_df, _ = process_adp(adp_files_data)
#         relay_df, _ = process_relay(relay_files_data)
#         dp_df = read_flexible_file(driverpay_data, dp_name)
#         ov_df = read_flexible_file(override_data, ov_name) if override_data else pd.DataFrame()

#         # Logic: Fuzzy Match
#         adp_list = adp_df['Driver'].unique().tolist()
#         relay_df['Driver'] = relay_df['Driver'].apply(lambda x: process.extractOne(x, adp_list, scorer=fuzz.token_sort_ratio)[0] if process.extractOne(x, adp_list, scorer=fuzz.token_sort_ratio)[1] >= 60 else x)
        
#         # Pivot
#         r_piv = relay_df.pivot_table(index='Driver', columns='Date', values='Hours', aggfunc='sum').fillna(0)
#         r_piv.columns = [f"R_{c}" for c in r_piv.columns]
#         a_piv = adp_df.pivot_table(index='Driver', columns='Date', values='Hours', aggfunc='sum').fillna(0)
#         a_piv.columns = [f"A_{c}" for c in a_piv.columns]
        
#         final = pd.concat([r_piv, a_piv], axis=1).fillna(0).reset_index()
#         dp_df['Drivers'] = dp_df['Drivers'].str.upper().str.strip()
#         final = final.merge(dp_df, left_on='Driver', right_on='Drivers', how='left')
#         final['Pay Type'] = final['Fixed Pay'].apply(lambda x: "FIXED" if x > 0 else "HOURLY")
        
#         ov_dict = {}
#         if not ov_df.empty:
#             ov_df['Driver'] = ov_df['Driver'].str.upper().str.strip()
#             ov_dict = {(r['Driver'], pd.to_datetime(r['Date']).date()): r['Override Price'] for _, r in ov_df.iterrows()}

#         excel_out = create_excel(final, list(r_piv.columns), list(a_piv.columns), ov_dict)
        
#         all_dates = [datetime.strptime(c.replace("A_", ""), "%Y-%m-%d").date() for c in a_piv.columns]
#         sheet_name = f"{min(all_dates).strftime('%d-%b')} to {max(all_dates).strftime('%d-%b')}" if all_dates else "Report"

#         if output_dest == "SharePoint":
#             existing_id = check_workbook_exists(st.session_state.access_token, st.session_state.site_info['id'], output_folder, "ADP.xlsx")[1]
#             if wb_action == "Add Sheet" and existing_id:
#                 old_bytes = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], existing_id)[0]
#                 updated, _ = add_sheet_to_workbook(old_bytes, excel_out.getvalue(), sheet_name)
#                 upload_to_sharepoint(st.session_state.access_token, st.session_state.site_info['id'], output_folder, "ADP.xlsx", updated.getvalue())
#             else:
#                 upload_to_sharepoint(st.session_state.access_token, st.session_state.site_info['id'], output_folder, "ADP.xlsx", excel_out.getvalue())
#             st.success("✅ Saved to SharePoint")
        
#         st.download_button("⬇️ Download Excel", excel_out, f"Payroll_{sheet_name}.xlsx")
#         st.dataframe(final)

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from rapidfuzz import process, fuzz
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from collections import Counter
import msal
import requests
from urllib.parse import urlencode, parse_qs, urlparse
import webbrowser

# ============ PAGE CONFIGURATION ============
st.set_page_config(
    page_title="Payroll Processor",
    page_icon="💰",
    layout="wide"
)

# ============ CONSTANTS ============
REG_RATE = 24
OT_RATE = 36

# Load Azure credentials from Streamlit secrets
try:
    CLIENT_ID = st.secrets["azure"]["client_id"]
    TENANT_ID = st.secrets["azure"]["tenant_id"]
    CLIENT_SECRET = st.secrets["azure"]["client_secret"]
    CREDENTIALS_CONFIGURED = True
except:
    CLIENT_ID = None
    TENANT_ID = None
    CLIENT_SECRET = None
    CREDENTIALS_CONFIGURED = False

# OAuth Configuration
REDIRECT_URI = "https://adppayroll.streamlit.app/" 
SCOPES = ["https://graph.microsoft.com/Files.ReadWrite.All",
          "https://graph.microsoft.com/Sites.ReadWrite.All", 
          "https://graph.microsoft.com/User.Read"]

# Excel Styles
HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
RELAY_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid') 
ADP_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  
DRIVER_FILL = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
OVERRIDE_FILL = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
ANOMALY_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ============ SESSION STATE ============
if 'access_token' not in st.session_state:
    st.session_state.access_token = None
if 'user_info' not in st.session_state:
    st.session_state.user_info = None
if 'site_info' not in st.session_state:
    st.session_state.site_info = None
if 'input_path' not in st.session_state:
    st.session_state.input_path = "root"
if 'output_path' not in st.session_state:
    st.session_state.output_path = "root"
if 'adp_files_data' not in st.session_state:
    st.session_state.adp_files_data = []
if 'relay_files_data' not in st.session_state:
    st.session_state.relay_files_data = []
if 'driverpay_data' not in st.session_state:
    st.session_state.driverpay_data = None
if 'override_data' not in st.session_state:
    st.session_state.override_data = None
if 'dp_name' not in st.session_state:
    st.session_state.dp_name = ""
if 'ov_name' not in st.session_state:
    st.session_state.ov_name = ""

# ============ HELPER FUNCTIONS ============

def read_flexible_file(content, filename):
    """Handles both CSV and Excel reading from binary content."""
    try:
        if filename.lower().endswith('.csv'):
            return pd.read_csv(io.BytesIO(content))
        elif filename.lower().endswith(('.xlsx', '.xls')):
            return pd.read_excel(io.BytesIO(content))
    except Exception as e:
        st.error(f"Error reading {filename}: {e}")
    return None

# ============ OAUTH & SHAREPOINT FUNCTIONS ============

def get_auth_url():
    auth_endpoint = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/authorize"
    params = {'client_id': CLIENT_ID, 'response_type': 'code', 'redirect_uri': REDIRECT_URI, 'scope': ' '.join(SCOPES), 'state': 'payroll'}
    return f"{auth_endpoint}?{urlencode(params)}"

def exchange_code_for_token(auth_code):
    try:
        app = msal.ConfidentialClientApplication(CLIENT_ID, client_credential=CLIENT_SECRET, authority=f"https://login.microsoftonline.com/{TENANT_ID}")
        result = app.acquire_token_by_authorization_code(auth_code, scopes=SCOPES, redirect_uri=REDIRECT_URI)
        return result.get("access_token"), result.get("error_description")
    except Exception as e: 
        return None, str(e)

def get_site_from_url(access_token, sharepoint_url):
    try:
        parsed = urlparse(sharepoint_url)
        path_parts = parsed.path.strip('/').split('/')
        headers = {'Authorization': f'Bearer {access_token}'}
        url = f"https://graph.microsoft.com/v1.0/sites/{parsed.hostname}:/sites/{path_parts[path_parts.index('sites')+1]}" if 'sites' in path_parts else f"https://graph.microsoft.com/v1.0/sites/{parsed.hostname}"
        resp = requests.get(url, headers=headers)
        return (resp.json(), None) if resp.status_code == 200 else (None, f"Error: {resp.status_code}")
    except Exception as e: 
        return None, str(e)

def list_sharepoint_files(access_token, site_id, path="root"):
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children" if path == "root" else f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{path}:/children"
        resp = requests.get(url, headers=headers)
        return (resp.json().get('value', []), None) if resp.status_code == 200 else (None, "Folder error")
    except Exception as e: 
        return None, str(e)

def download_sharepoint_file(access_token, site_id, file_id):
    try:
        resp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/content", headers={'Authorization': f'Bearer {access_token}'})
        return (resp.content, None) if resp.status_code == 200 else (None, "Download error")
    except Exception as e: 
        return None, str(e)

def upload_to_sharepoint(access_token, site_id, folder_path, filename, file_content):
    try:
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}/{filename}:/content" if folder_path != "root" else f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{filename}:/content"
        resp = requests.put(url, headers={'Authorization': f'Bearer {access_token}'}, data=file_content)
        return (True, None) if resp.status_code in [200, 201] else (False, f"Error: {resp.status_code}")
    except Exception as e: 
        return False, str(e)

def check_workbook_exists(access_token, site_id, folder_path, workbook_name="ADP.xlsx"):
    try:
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}/{workbook_name}" if folder_path != "root" else f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{workbook_name}"
        resp = requests.get(url, headers={'Authorization': f'Bearer {access_token}'})
        return (True, resp.json().get('id')) if resp.status_code == 200 else (False, None)
    except: 
        return False, None

def add_sheet_to_workbook(existing_wb_bytes, new_sheet_bytes, sheet_name):
    try:
        wb = load_workbook(io.BytesIO(existing_wb_bytes))
        if sheet_name in wb.sheetnames: 
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        new_wb = load_workbook(io.BytesIO(new_sheet_bytes))
        new_ws = new_wb.active
        for row in new_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font, new_cell.border = cell.font.copy(), cell.border.copy()
                    new_cell.fill, new_cell.number_format = cell.fill.copy(), cell.number_format
                    new_cell.alignment = cell.alignment.copy()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, None
    except Exception as e: 
        return None, str(e)

# ============ AUTH CALLBACK HANDLING ============
if 'code' in st.query_params and not st.session_state.access_token:
    token, err = exchange_code_for_token(st.query_params['code'])
    if token:
        st.session_state.access_token = token
        resp = requests.get('https://graph.microsoft.com/v1.0/me', headers={'Authorization': f'Bearer {token}'})
        if resp.status_code == 200: 
            st.session_state.user_info = resp.json()
    st.query_params.clear()
    st.rerun()

# ============ UI: HEADER ============
st.title("🚛 Driver Payroll Processing System")
st.markdown("---")

# ============ UI: SIDEBAR ============
with st.sidebar:
    st.header("🔐 Microsoft Login")
    
    if not st.session_state.access_token:
        if st.button("🔑 Login with Microsoft", type="primary"):
            st.markdown(f"### [Click here to login]({get_auth_url()})")
            st.info("Sign in with your Microsoft email and password")
            st.stop()
    else:
        st.success(f"✅ {st.session_state.user_info.get('displayName', 'User')}")
        st.caption(f"📧 {st.session_state.user_info.get('mail', st.session_state.user_info.get('userPrincipalName', ''))}")
        
        if st.button("🚪 Logout"):
            for key in list(st.session_state.keys()): 
                del st.session_state[key]
            st.rerun()
    
    st.markdown("---")
    
    # SharePoint Connection
    if st.session_state.access_token:
        st.header("📁 SharePoint Connection")
        sh_url = st.text_input("SharePoint Site URL", placeholder="https://company.sharepoint.com/sites/YourSite")
        
        if st.button("🔗 Connect") and sh_url:
            site, err = get_site_from_url(st.session_state.access_token, sh_url)
            if site:
                st.session_state.site_info = site
                st.success(f"Connected: {site.get('displayName', 'Site')}")
            else:
                st.error(f"Connection failed: {err}")
        
        if st.session_state.site_info:
            st.success(f"📍 {st.session_state.site_info.get('displayName', 'Connected')}")

# ============ FILE SELECTION WITH INDIVIDUAL SOURCE OPTIONS ============

st.subheader("📂 Input Files")

# ADP Files
with st.expander("📊 ADP Files", expanded=True):
    adp_source = st.radio("ADP Source:", ["Desktop", "SharePoint"], key="adp_source", horizontal=True)
    
    if adp_source == "Desktop":
        adp_uploads = st.file_uploader("Upload ADP CSV files", type=['csv'], accept_multiple_files=True, key="adp_up")
        if adp_uploads:
            st.session_state.adp_files_data = [f.read() for f in adp_uploads]
            st.success(f"✅ {len(adp_uploads)} ADP files loaded")
    
    else:  # SharePoint
        if st.session_state.site_info:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.caption(f"📁 Path: {st.session_state.input_path}")
            with col2:
                if st.button("⬆️ Up", key="adp_up_btn") and st.session_state.input_path != "root":
                    st.session_state.input_path = "/".join(st.session_state.input_path.split('/')[:-1]) if '/' in st.session_state.input_path else "root"
                    st.rerun()
            
            items, _ = list_sharepoint_files(st.session_state.access_token, st.session_state.site_info['id'], st.session_state.input_path)
            
            if items:
                folders = [i for i in items if 'folder' in i]
                csv_files = [i for i in items if i['name'].lower().endswith('.csv')]
                
                if folders:
                    folder_sel = st.selectbox("📂 Navigate to folder:", ["-- Stay Here --"] + [f['name'] for f in folders], key="adp_folder")
                    if folder_sel != "-- Stay Here --" and st.button("Enter Folder", key="adp_enter"):
                        st.session_state.input_path = folder_sel if st.session_state.input_path == "root" else f"{st.session_state.input_path}/{folder_sel}"
                        st.rerun()
                
                if csv_files:
                    adp_sel = st.multiselect("Select ADP CSV files:", [f['name'] for f in csv_files], key="adp_sel")
                    
                    if st.button("Confirm ADP Files", key="adp_confirm"):
                        with st.spinner("Downloading ADP files..."):
                            st.session_state.adp_files_data = []
                            for filename in adp_sel:
                                file_obj = next(f for f in csv_files if f['name'] == filename)
                                content, err = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], file_obj['id'])
                                if content:
                                    st.session_state.adp_files_data.append(content)
                            st.success(f"✅ {len(st.session_state.adp_files_data)} ADP files downloaded")
                else:
                    st.info("No CSV files in current folder")
        else:
            st.warning("⚠️ Connect to SharePoint first")

# Relay Files
with st.expander("🚚 Relay Files", expanded=True):
    relay_source = st.radio("Relay Source:", ["Desktop", "SharePoint"], key="relay_source", horizontal=True)
    
    if relay_source == "Desktop":
        relay_uploads = st.file_uploader("Upload Relay CSV files", type=['csv'], accept_multiple_files=True, key="relay_up")
        if relay_uploads:
            st.session_state.relay_files_data = [f.read() for f in relay_uploads]
            st.success(f"✅ {len(relay_uploads)} Relay files loaded")
    
    else:  # SharePoint
        if st.session_state.site_info:
            items, _ = list_sharepoint_files(st.session_state.access_token, st.session_state.site_info['id'], st.session_state.input_path)
            
            if items:
                csv_files = [i for i in items if i['name'].lower().endswith('.csv')]
                
                if csv_files:
                    relay_sel = st.multiselect("Select Relay CSV files:", [f['name'] for f in csv_files], key="relay_sel")
                    
                    if st.button("Confirm Relay Files", key="relay_confirm"):
                        with st.spinner("Downloading Relay files..."):
                            st.session_state.relay_files_data = []
                            for filename in relay_sel:
                                file_obj = next(f for f in csv_files if f['name'] == filename)
                                content, err = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], file_obj['id'])
                                if content:
                                    st.session_state.relay_files_data.append(content)
                            st.success(f"✅ {len(st.session_state.relay_files_data)} Relay files downloaded")
                else:
                    st.info("No CSV files in current folder")
        else:
            st.warning("⚠️ Connect to SharePoint first")

# DriverPay File
with st.expander("💵 DriverPay File", expanded=True):
    dp_source = st.radio("DriverPay Source:", ["Desktop", "SharePoint"], key="dp_source", horizontal=True)
    
    if dp_source == "Desktop":
        dp_upload = st.file_uploader("Upload DriverPay file", type=['csv', 'xlsx'], key="dp_up")
        if dp_upload:
            st.session_state.driverpay_data = dp_upload.read()
            st.session_state.dp_name = dp_upload.name
            st.success(f"✅ {dp_upload.name} loaded")
    
    else:  # SharePoint
        if st.session_state.site_info:
            items, _ = list_sharepoint_files(st.session_state.access_token, st.session_state.site_info['id'], st.session_state.input_path)
            
            if items:
                flex_files = [i for i in items if i['name'].lower().endswith(('.csv', '.xlsx', '.xls'))]
                
                if flex_files:
                    dp_sel = st.selectbox("Select DriverPay file:", ["-- None --"] + [f['name'] for f in flex_files], key="dp_sel")
                    
                    if dp_sel != "-- None --" and st.button("Confirm DriverPay File", key="dp_confirm"):
                        with st.spinner("Downloading DriverPay file..."):
                            file_obj = next(f for f in flex_files if f['name'] == dp_sel)
                            content, err = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], file_obj['id'])
                            if content:
                                st.session_state.driverpay_data = content
                                st.session_state.dp_name = file_obj['name']
                                st.success(f"✅ {file_obj['name']} downloaded")
                else:
                    st.info("No CSV/Excel files in current folder")
        else:
            st.warning("⚠️ Connect to SharePoint first")

# Override File (Optional)
with st.expander("🔶 Override File (Optional)", expanded=False):
    ov_source = st.radio("Override Source:", ["Desktop", "SharePoint"], key="ov_source", horizontal=True)
    
    if ov_source == "Desktop":
        ov_upload = st.file_uploader("Upload Override file", type=['csv', 'xlsx'], key="ov_up")
        if ov_upload:
            st.session_state.override_data = ov_upload.read()
            st.session_state.ov_name = ov_upload.name
            st.success(f"✅ {ov_upload.name} loaded")
    
    else:  # SharePoint
        if st.session_state.site_info:
            items, _ = list_sharepoint_files(st.session_state.access_token, st.session_state.site_info['id'], st.session_state.input_path)
            
            if items:
                flex_files = [i for i in items if i['name'].lower().endswith(('.csv', '.xlsx', '.xls'))]
                
                if flex_files:
                    ov_sel = st.selectbox("Select Override file:", ["-- None --"] + [f['name'] for f in flex_files], key="ov_sel")
                    
                    if ov_sel != "-- None --" and st.button("Confirm Override File", key="ov_confirm"):
                        with st.spinner("Downloading Override file..."):
                            file_obj = next(f for f in flex_files if f['name'] == ov_sel)
                            content, err = download_sharepoint_file(st.session_state.access_token, st.session_state.site_info['id'], file_obj['id'])
                            if content:
                                st.session_state.override_data = content
                                st.session_state.ov_name = file_obj['name']
                                st.success(f"✅ {file_obj['name']} downloaded")
                else:
                    st.info("No CSV/Excel files in current folder")
        else:
            st.warning("⚠️ Connect to SharePoint first")

# ============ OUTPUT CONFIGURATION ============
st.markdown("---")
st.subheader("📤 Output Configuration")

output_dest = st.radio("Save to:", ["Download Only", "SharePoint"], horizontal=True)

if output_dest == "SharePoint" and st.session_state.site_info:
    st.markdown("### Select Output Folder")
    
    col1, col2 = st.columns([3, 1])
    with col1:
        st.caption(f"📁 Output Path: {st.session_state.output_path}")
    with col2:
        if st.button("⬆️ Up", key="output_up") and st.session_state.output_path != "root":
            st.session_state.output_path = "/".join(st.session_state.output_path.split('/')[:-1]) if '/' in st.session_state.output_path else "root"
            st.rerun()
    
    items, _ = list_sharepoint_files(st.session_state.access_token, st.session_state.site_info['id'], st.session_state.output_path)
    
    if items:
        folders = [i for i in items if 'folder' in i]
        
        if folders:
            folder_sel = st.selectbox("📂 Navigate to output folder:", ["-- Use Current --"] + [f['name'] for f in folders], key="output_folder")
            if folder_sel != "-- Use Current --" and st.button("Enter Folder", key="output_enter"):
                st.session_state.output_path = folder_sel if st.session_state.output_path == "root" else f"{st.session_state.output_path}/{folder_sel}"
                st.rerun()
    
    wb_action = st.radio("Workbook Action:", ["Add Sheet to Existing ADP.xlsx", "Create New Workbook"], horizontal=True)

# ============ CORE PROCESSING LOGIC ============

def process_adp(contents):
    dfs = [pd.read_csv(io.BytesIO(c)) for c in contents]
    df = pd.concat(dfs)
    df['Driver'] = df['Payroll Name'].astype(str).str.replace(',', '').str.strip().str.upper()
    df['Date'] = pd.to_datetime(df['Pay Date']).dt.date
    df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce').fillna(0)
    return df.groupby(['Driver', 'Date'], as_index=False)['Hours'].sum(), None

def process_relay(contents):
    dfs = [pd.read_csv(io.BytesIO(c)) for c in contents]
    df = pd.concat(dfs)
    df['Stop1_Actual'] = pd.to_datetime(df['Stop 1  Actual Arrival Date'] + " " + df['Stop 1  Actual Arrival Time'], errors='coerce')
    df['Stop2_Actual'] = pd.to_datetime(df['Stop 2  Actual Arrival Date'] + " " + df['Stop 2  Actual Arrival Time'], errors='coerce')
    df['Stop1_Planned'] = pd.to_datetime(df['Stop 1 Planned Arrival Date'] + " " + df['Stop 1 Planned Arrival Time'], errors='coerce')
    df = df.dropna(subset=['Stop1_Actual', 'Stop2_Actual'])
    rows = []
    for tid, gp in df.groupby('Trip ID'):
        gp = gp.sort_values('Stop1_Actual')
        s_act, s_pln, e_act = gp.iloc[0]['Stop1_Actual'], gp.iloc[0]['Stop1_Planned'], gp.iloc[-1]['Stop2_Actual']
        if e_act < s_act: e_act += pd.Timedelta(days=1)
        raw = (e_act - s_act).total_seconds() / 3600
        start = s_pln if raw > 20 and pd.notna(s_pln) else s_act
        rows.append([gp['Driver Name'].str.upper().iloc[0], start.date(), round((e_act-start).total_seconds()/3600, 2)])
    return pd.DataFrame(rows, columns=['Driver', 'Date', 'Hours']), None

def create_excel(final_df, r_cols, a_cols, ov_dict):
    """Create Excel with full formatting and formulas (same as your original)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    
    calcs = ["Total Relay Hours", "Relay_Loads", "W1 Hours", "W1 Regular", "W1 OT", 
             "W2 Hours", "W2 Regular", "W2 OT", "Total ADP Hours", "Total Regular", "Total OT", 
             "Override Pay", "Final Pay", "Equivalent Hours", "Hour Adjustment"]
    
    r_fmt = [datetime.strptime(c.replace("R_", ""), "%Y-%m-%d").date().strftime("%d-%b") for c in r_cols]
    a_fmt = [datetime.strptime(c.replace("A_", ""), "%Y-%m-%d").date().strftime("%d-%b") for c in a_cols]
    
    headers = [c for c in final_df.columns if not c.startswith(('R_','A_'))] + r_fmt + a_fmt + calcs
    ws.append(headers)
    
    col_m = {n: get_column_letter(i) for i, n in enumerate(headers, 1)}
    override_cells, anomaly_rows = [], []
    
    for r_idx, row_vals in enumerate(final_df.values, start=2):
        row_n = str(r_idx)
        driver = row_vals[0]
        col_ptr = 1
        
        # Write base columns
        for i, val in enumerate(row_vals):
            if not final_df.columns[i].startswith(('R_','A_')):
                ws.cell(r_idx, col_ptr, val)
                col_ptr += 1
        
        # Write relay and ADP data
        has_relay, has_adp = False, False
        for c in r_cols:
            val = final_df.loc[r_idx-2, c]
            ws.cell(r_idx, col_ptr, val)
            if val > 0: has_relay = True
            col_ptr += 1
        
        for c in a_cols:
            val = final_df.loc[r_idx-2, c]
            ws.cell(r_idx, col_ptr, val)
            if val > 0: has_adp = True
            
            # Check for override
            date_str = c.replace("A_", "")
            try:
                work_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                if (driver, work_date) in ov_dict:
                    override_cells.append(f"{get_column_letter(col_ptr)}{r_idx}")
            except:
                pass
            col_ptr += 1
        
        if has_relay and not has_adp:
            anomaly_rows.append(r_idx)
        
        # Formulas
        if r_fmt:
            ws.cell(r_idx, col_ptr).value = f"=SUM({col_m[r_fmt[0]]}{row_n}:{col_m[r_fmt[-1]]}{row_n})"
            ws.cell(r_idx, col_ptr+1).value = f'=COUNTIF({col_m[r_fmt[0]]}{row_n}:{col_m[r_fmt[-1]]}{row_n}, ">0")'
        col_ptr += 2
        
        # W1/W2 calculations
        if len(a_fmt) >= 7:
            ws.cell(r_idx, col_ptr).value = f"=SUM({col_m[a_fmt[0]]}{row_n}:{col_m[a_fmt[6]]}{row_n})"
            ws.cell(r_idx, col_ptr+1).value = f"=MIN({col_m['W1 Hours']}{row_n}, 40)"
            ws.cell(r_idx, col_ptr+2).value = f"=MAX(0, {col_m['W1 Hours']}{row_n}-40)"
        col_ptr += 3
        
        if len(a_fmt) >= 14:
            ws.cell(r_idx, col_ptr).value = f"=SUM({col_m[a_fmt[7]]}{row_n}:{col_m[a_fmt[13]]}{row_n})"
            ws.cell(r_idx, col_ptr+1).value = f"=MIN({col_m['W2 Hours']}{row_n}, 40)"
            ws.cell(r_idx, col_ptr+2).value = f"=MAX(0, {col_m['W2 Hours']}{row_n}-40)"
        col_ptr += 3
        
        # Totals
        ws.cell(r_idx, col_ptr).value = f"={col_m['W1 Hours']}{row_n}+{col_m['W2 Hours']}{row_n}"
        ws.cell(r_idx, col_ptr+1).value = f"={col_m['W1 Regular']}{row_n}+{col_m['W2 Regular']}{row_n}"
        ws.cell(r_idx, col_ptr+2).value = f"={col_m['W1 OT']}{row_n}+{col_m['W2 OT']}{row_n}"
        col_ptr += 3
        
        # Override pay
        override_total = sum(v for (d, dt), v in ov_dict.items() if d == driver)
        ws.cell(r_idx, col_ptr).value = override_total
        col_ptr += 1
        
        # Final pay formula
        pt, fixed = col_m["Pay Type"], col_m["Fixed Pay"]
        target = col_m.get("DriverPay_Target_Loads", "1")
        actual = col_m["Relay_Loads"]
        tot_adp = col_m["Total ADP Hours"]
        w1_ot, w2_ot = col_m["W1 OT"], col_m["W2 OT"]
        ov_pay = col_m["Override Pay"]
        
        ws.cell(r_idx, col_ptr).value = (
            f"=IF({pt}{row_n}=\"FIXED\", "
            f"({fixed}{row_n}/MAX(1,{target}{row_n}))*{actual}{row_n}, "
            f"({tot_adp}{row_n}*{REG_RATE})+({w1_ot}{row_n}*{OT_RATE-REG_RATE})+({w2_ot}{row_n}*{OT_RATE-REG_RATE}))"
            f"+{ov_pay}{row_n}"
        )
        col_ptr += 1
        
        # Equivalent hours and adjustment
        f_pay = col_m["Final Pay"]
        ws.cell(r_idx, col_ptr).value = f"={f_pay}{row_n}/{REG_RATE}"
        ws.cell(r_idx, col_ptr+1).value = f"={col_m['Equivalent Hours']}{row_n}-{tot_adp}{row_n}"
    
    # Styling
    adj_letter = col_m["Hour Adjustment"]
    ws.conditional_formatting.add(f'{adj_letter}2:{adj_letter}{ws.max_row}', 
                                 CellIsRule(operator='lessThan', formula=['0'], fill=RED_FILL))
    ws.conditional_formatting.add(f'{adj_letter}2:{adj_letter}{ws.max_row}', 
                                 CellIsRule(operator='greaterThan', formula=['0'], fill=YELLOW_FILL))
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        for cell in row:
            header = headers[cell.column-1]
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            
            if r_idx > 1 and header not in ["Driver", "Pay Type"]:
                cell.number_format = '[=0]"";#,##0.00'
            
            if r_idx == 1:
                cell.fill, cell.font = HEADER_FILL, Font(bold=True)
            else:
                if r_idx in anomaly_rows and header == "Driver":
                    cell.fill, cell.font = ANOMALY_FILL, Font(bold=True, color="FFFFFF")
                elif f"{get_column_letter(cell.column)}{cell.row}" in override_cells:
                    cell.fill, cell.font = OVERRIDE_FILL, Font(bold=True, color="FFFFFF")
                elif header == "Driver":
                    cell.fill = DRIVER_FILL
                elif header in r_fmt or "Relay" in header:
                    cell.fill = RELAY_FILL
                elif header in a_fmt or "ADP" in header or "W" in header or "OT" in header or "Total" in header or "Regular" in header:
                    cell.fill = ADP_FILL
                elif header == "Override Pay" and cell.value and cell.value > 0:
                    cell.fill, cell.font = OVERRIDE_FILL, Font(bold=True, color="FFFFFF")
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 13
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, anomaly_rows

# ============ PROCESS BUTTON ============
st.markdown("---")

if st.button("🚀 Process Payroll", type="primary", use_container_width=True):
    # Validation
    if not st.session_state.adp_files_data:
        st.error("❌ Missing ADP files")
        st.stop()
    
    if not st.session_state.relay_files_data:
        st.error("❌ Missing Relay files")
        st.stop()
    
    if not st.session_state.driverpay_data:
        st.error("❌ Missing DriverPay file")
        st.stop()
    
    with st.spinner("Processing payroll data..."):
        # Process files
        adp_df, _ = process_adp(st.session_state.adp_files_data)
        relay_df, _ = process_relay(st.session_state.relay_files_data)
        dp_df = read_flexible_file(st.session_state.driverpay_data, st.session_state.dp_name)
        
        # Override
        ov_dict = {}
        if st.session_state.override_data:
            ov_df = read_flexible_file(st.session_state.override_data, st.session_state.ov_name)
            if ov_df is not None and not ov_df.empty:
                ov_df['Driver'] = ov_df['Driver'].str.upper().str.strip()
                ov_df['Date'] = pd.to_datetime(ov_df['Date']).dt.date
                ov_dict = {(r['Driver'], r['Date']): r['Override Price'] for _, r in ov_df.iterrows()}
        
        # Fuzzy matching
        adp_list = adp_df['Driver'].unique().tolist()
        relay_df['Driver'] = relay_df['Driver'].apply(
            lambda x: process.extractOne(x, adp_list, scorer=fuzz.token_sort_ratio)[0] 
            if process.extractOne(x, adp_list, scorer=fuzz.token_sort_ratio)[1] >= 60 else x
        )
        
        # Pivot
        r_piv = relay_df.pivot_table(index='Driver', columns='Date', values='Hours', aggfunc='sum').fillna(0)
        if len(r_piv.columns) > 1:
            r_piv = r_piv.iloc[:, 1:]
        r_piv.columns = [f"R_{c}" for c in r_piv.columns]
        
        a_piv = adp_df.pivot_table(index='Driver', columns='Date', values='Hours', aggfunc='sum').fillna(0)
        a_piv.columns = [f"A_{c}" for c in a_piv.columns]
        
        # Merge
        final = pd.concat([r_piv, a_piv], axis=1).fillna(0).reset_index()
        dp_df['Drivers'] = dp_df['Drivers'].str.upper().str.strip()
        
        if 'Fixed Pay' in dp_df.columns:
            dp_df['Fixed Pay'] = pd.to_numeric(
                dp_df['Fixed Pay'].astype(str).str.replace(r'[\$,]', '', regex=True), 
                errors='coerce'
            ).fillna(0)
        
        if 'Total Loads' in dp_df.columns:
            dp_df = dp_df.rename(columns={'Total Loads': 'DriverPay_Target_Loads'})
        
        final = final.merge(dp_df, left_on='Driver', right_on='Drivers', how='left')
        final = final.drop(columns=[c for c in ['Drivers', 'Unnamed: 3', 'Unnamed: 4'] if c in final.columns])
        final['Pay Type'] = final['Fixed Pay'].apply(lambda x: "FIXED" if pd.notna(x) and x > 0 else "HOURLY")
        
        # Reorder columns
        base_cols = ['Driver', 'Pay Type', 'Fixed Pay']
        if 'DriverPay_Target_Loads' in final.columns:
            base_cols.append('DriverPay_Target_Loads')
        other_cols = [c for c in final.columns if c not in base_cols and not c.startswith(('R_', 'A_'))]
        base_cols.extend(other_cols)
        
        r_cols = [c for c in final.columns if c.startswith("R_")]
        a_cols = [c for c in final.columns if c.startswith("A_")]
        final = final[base_cols + r_cols + a_cols]
        
        # Generate Excel
        excel_out, anomaly_rows = create_excel(final, r_cols, a_cols, ov_dict)
        
        # Sheet name
        all_dates = [datetime.strptime(c.replace("A_", ""), "%Y-%m-%d").date() for c in a_cols]
        sheet_name = f"{min(all_dates).strftime('%d-%b')} to {max(all_dates).strftime('%d-%b')}" if all_dates else "Payroll"
        
        st.success(f"✅ Excel generated! Sheet: **{sheet_name}**")
        
        # Save to SharePoint if requested
        if output_dest == "SharePoint" and st.session_state.site_info:
            wb_exists, file_id = check_workbook_exists(
                st.session_state.access_token, 
                st.session_state.site_info['id'], 
                st.session_state.output_path, 
                "ADP.xlsx"
            )
            
            if "Add Sheet" in wb_action and wb_exists:
                st.info("Adding sheet to existing workbook...")
                old_bytes, _ = download_sharepoint_file(
                    st.session_state.access_token, 
                    st.session_state.site_info['id'], 
                    file_id
                )
                if old_bytes:
                    updated, _ = add_sheet_to_workbook(old_bytes, excel_out.getvalue(), sheet_name)
                    if updated:
                        success, _ = upload_to_sharepoint(
                            st.session_state.access_token, 
                            st.session_state.site_info['id'], 
                            st.session_state.output_path, 
                            "ADP.xlsx", 
                            updated.getvalue()
                        )
                        if success:
                            st.success("✅ Uploaded to SharePoint!")
            else:
                st.info("Creating new workbook...")
                # Rename the sheet in new workbook
                wb_temp = load_workbook(excel_out)
                wb_temp.active.title = sheet_name
                temp_out = io.BytesIO()
                wb_temp.save(temp_out)
                temp_out.seek(0)
                
                success, _ = upload_to_sharepoint(
                    st.session_state.access_token, 
                    st.session_state.site_info['id'], 
                    st.session_state.output_path, 
                    "ADP.xlsx", 
                    temp_out.getvalue()
                )
                if success:
                    st.success("✅ Created ADP.xlsx in SharePoint!")
        
        # Show warnings
        if anomaly_rows:
            st.error(f"🚨 {len(anomaly_rows)} anomalies detected (Relay without ADP) - marked in RED")
        
        if ov_dict:
            st.warning(f"🔶 {len(ov_dict)} override payments applied - highlighted in ORANGE")
        
        # Preview
        st.subheader("📊 Preview")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Drivers", len(final))
        col2.metric("Relay Days", len(r_cols))
        col3.metric("ADP Days", len(a_cols))
        col4.metric("Overrides", len(ov_dict))
        
        st.dataframe(final, use_container_width=True, height=400)
        
        # Download button
        st.download_button(
            "⬇️ Download Excel", 
            excel_out.getvalue(), 
            f"Payroll_{sheet_name}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

st.markdown("---")
st.caption("💡 Login → Connect SharePoint → Select Files → Process Payroll")